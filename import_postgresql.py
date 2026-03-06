"""
Import des données Excel vers PostgreSQL
Crimes et Délits 2012-2021 (PN + GN)
"""

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
import re, os, time

# ── Config ────────────────────────────────────────────────
PG = dict(host="localhost", port=5432, dbname="crimes_delits",
          user="postgres", password="user")
EXCEL = "crimes-et-delits-enregistres-par-les-services-de-gendarmerie-et-de-police-depuis-2012.xlsx"
BATCH = 5000   # lignes insérées par transaction
# ──────────────────────────────────────────────────────────

def connect():
    return psycopg2.connect(**PG)

# ─── Insertion des forces ─────────────────────────────────

def insert_forces(cur):
    execute_values(cur,
        "INSERT INTO force (id_force, code, libelle) VALUES %s ON CONFLICT DO NOTHING",
        [(1, "PN", "Police Nationale"),
         (2, "GN", "Gendarmerie Nationale")])

# ─── Lecture d'une feuille Excel ─────────────────────────

def parse_sheet(ws, force_code):
    rows   = list(ws.iter_rows(values_only=True))
    is_pn  = (force_code == "PN")
    row_dept  = rows[0]
    row_peri  = rows[1] if is_pn else None
    row_svc   = rows[2] if is_pn else rows[1]
    data_start= 3       if is_pn else 2

    services = []
    for col in range(2, len(row_svc)):
        nom = row_svc[col]
        if nom is None:
            break
        code_dept = str(row_dept[col]) if row_dept[col] is not None else "??"
        code_dir  = str(row_peri[col]) if (is_pn and row_peri and row_peri[col]) else None
        services.append((nom, code_dept, code_dir))

    types_faits = {}
    stats       = []
    for row in rows[data_start:]:
        code_index = row[0]
        libelle    = row[1]
        if not isinstance(code_index, int):
            continue
        if libelle and "non utilis" in str(libelle).lower():
            continue
        types_faits[code_index] = libelle
        for s_idx, _ in enumerate(services):
            val    = row[s_idx + 2]
            nombre = int(val) if isinstance(val, (int, float)) and val else 0
            if nombre > 0:
                stats.append((s_idx, code_index, nombre))

    return {"types_faits": types_faits, "services": services, "stats": stats}

# ─── Import principal ─────────────────────────────────────

def import_excel(conn):
    cur = conn.cursor()

    # Forces
    insert_forces(cur)
    conn.commit()

    # Pré-collecte des directions (PN uniquement)
    print("[...] Lecture des directions...")
    wb   = openpyxl.load_workbook(EXCEL, read_only=True)
    dirs = set()
    for name in wb.sheetnames:
        if "PN" in name:
            ws   = wb[name]
            row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            for v in row2[2:]:
                if v:
                    dirs.add(str(v))
    if dirs:
        execute_values(cur,
            "INSERT INTO direction (code, libelle) VALUES %s ON CONFLICT DO NOTHING",
            [(d, d) for d in dirs])
        conn.commit()
    print(f"[OK] {len(dirs)} directions")

    # Cache direction code→id
    cur.execute("SELECT code, id_direction FROM direction")
    dir_cache = {row[0]: row[1] for row in cur.fetchall()}

    # Cache service existants (pour éviter les doublons inter-années)
    svc_cache = {}  # (nom, code_dept, id_force) → id_service

    t_total = time.perf_counter()

    for sheet_name in wb.sheetnames:
        if sheet_name in ("Présentation", "Presentation"):
            continue
        force_code = "PN" if "PN" in sheet_name else ("GN" if "GN" in sheet_name else None)
        if not force_code:
            continue
        m = re.search(r"(\d{4})", sheet_name)
        if not m:
            continue
        annee    = int(m.group(1))
        id_force = 1 if force_code == "PN" else 2

        t0 = time.perf_counter()
        print(f"[...] {sheet_name} ...", end=" ", flush=True)

        ws   = wb[sheet_name]
        data = parse_sheet(ws, force_code)

        # ── Types de faits
        if data["types_faits"]:
            execute_values(cur,
                "INSERT INTO type_fait (code_index, libelle) VALUES %s ON CONFLICT DO NOTHING",
                list(data["types_faits"].items()))

        # ── Départements
        depts = {svc[1] for svc in data["services"]}
        if depts:
            execute_values(cur,
                "INSERT INTO departement (code_dept) VALUES %s ON CONFLICT DO NOTHING",
                [(d,) for d in depts])

        # ── Services
        svc_id_map = {}
        new_services = []
        for s_idx, (nom, code_dept, code_dir) in enumerate(data["services"]):
            key = (nom, code_dept, id_force)
            if key in svc_cache:
                svc_id_map[s_idx] = svc_cache[key]
            else:
                id_dir = dir_cache.get(code_dir) if code_dir else None
                new_services.append((s_idx, nom, code_dept, id_force, id_dir))

        # Insertion batch des nouveaux services
        if new_services:
            for s_idx, nom, code_dept, idf, id_dir in new_services:
                cur.execute(
                    "INSERT INTO service (nom, code_dept, id_force, id_direction) "
                    "VALUES (%s,%s,%s,%s) RETURNING id_service",
                    (nom, code_dept, idf, id_dir))
                sid = cur.fetchone()[0]
                key = (nom, code_dept, id_force)
                svc_cache[key]    = sid
                svc_id_map[s_idx] = sid

        # ── Statistiques (par batch)
        batch = [
            (svc_id_map[s_idx], annee, code_index, nombre)
            for s_idx, code_index, nombre in data["stats"]
            if s_idx in svc_id_map
        ]
        for i in range(0, len(batch), BATCH):
            execute_values(cur,
                """INSERT INTO statistique (id_service, annee, code_index, nombre)
                   VALUES %s
                   ON CONFLICT (id_service, annee, code_index)
                   DO UPDATE SET nombre = EXCLUDED.nombre""",
                batch[i:i+BATCH])

        conn.commit()
        dt = time.perf_counter() - t0
        print(f"{len(data['services'])} services, {len(batch)} stats — {dt:.1f}s")

    dt_total = time.perf_counter() - t_total
    print(f"\n[DONE] Import total : {dt_total:.1f}s")

# ─── Résumé ───────────────────────────────────────────────

def print_resume(conn):
    cur = conn.cursor()
    print("\n=== RÉSUMÉ DE LA BASE POSTGRESQL ===")
    tables = ["force","direction","departement","service","type_fait","statistique"]
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        print(f"  {t:15s} : {cur.fetchone()[0]:>10,} lignes")

    print("\n--- Top 10 crimes (cumul 2012-2021) ---")
    cur.execute("""
        SELECT tf.libelle, SUM(st.nombre) AS total
        FROM statistique st
        JOIN type_fait tf ON tf.code_index = st.code_index
        GROUP BY tf.code_index, tf.libelle
        ORDER BY total DESC LIMIT 10
    """)
    for lib, total in cur.fetchall():
        print(f"  {total:>12,}  {lib}")

    print("\n--- Faits par année ---")
    cur.execute("""
        SELECT annee, SUM(nombre) FROM statistique
        GROUP BY annee ORDER BY annee
    """)
    for annee, total in cur.fetchall():
        print(f"  {annee} : {total:>12,}")

# ─── MAIN ─────────────────────────────────────────────────

if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    conn = connect()
    try:
        import_excel(conn)
        print_resume(conn)
    finally:
        conn.close()
